import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import tensorflow as tf
from sklearn.metrics import accuracy_score, f1_score, precision_score, recall_score, confusion_matrix, roc_auc_score, roc_curve

import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image

def plot_loss_graph(history, f_num=None):
    plt.plot(history.history['loss'], label='train loss')
    plt.plot(history.history['val_loss'], label='valid loss')
    plt.legend()
    plt.xlabel('Epoch'); plt.ylabel('loss')
    plt.savefig(f"../plot/loss_graph{f_num}.jpg")
    plt.close()

def plot_roc_curve(y_test, y_prob, f_num=None): 
    fpr , tpr , thresholds = roc_curve(y_test ,y_prob)
    auc_score = roc_auc_score(y_test, y_prob)
    plt.plot(fpr,tpr, label=f'AUC = {auc_score}')
    plt.axis([0,1,0,1])
    plt.plot([0, 1], [0, 1], 'k--')
    plt.xlabel('False Positive Rate') 
    plt.ylabel('True Positive Rate') 
    plt.legend()
    plt.savefig(f"../plot/roc_curve{f_num}.jpg")
    plt.close()

def plot_confusion_matrix(y_test, y_pred, classes=['ACC', 'REJ'], normalize=False, cmap=plt.cm.OrRd):
    cm = confusion_matrix(y_test, y_pred)
    
    title = ""
    if normalize:
        title = 'Normalized confusion matrix'
    else:
        title = 'Confusion matrix'
    
    # classes = classes[unique_labels(y_true, y_pred)]
    if normalize:
        # 정규화 할 때는 모든 값을 더해서 합이 1이 되도록 각 데이터를 스케일링 합니다.
        cm = cm.astype('float') / cm.sum(axis=1)[:, np.newaxis]
    # print(title, ":\n", cm)

    fig, ax = plt.subplots()
    im = ax.imshow(cm, interpolation='nearest', cmap=cmap)
    ax.figure.colorbar(im, ax=ax)
    ax.set(xticks=np.arange(cm.shape[1]),
           yticks=np.arange(cm.shape[0]),
           xticklabels=classes, yticklabels=classes,
           title=title,
           ylabel='True label',
           xlabel='Predicted label')

    # label을 45도 회전해서 보여주도록 변경
    # plt.setp(ax.get_xticklabels(), rotation=45, ha="right",
    #          rotation_mode="anchor")

    # confusion matrix 실제 값 뿌리기
    fmt = '.2f' if normalize else 'd'
    thresh = cm.max() / 2.
    for i in range(cm.shape[0]):
        for j in range(cm.shape[1]):
            ax.text(j, i, format(cm[i, j], fmt),
                    ha="center", va="center",
                    color="white" if cm[i, j] > thresh else "black")
    fig.tight_layout()
    
    plt.savefig('../plot/confusion matrix.png')
    plt.close()



def evaluate_data(y_test, y_pred):

    TP = 0
    FP = 0
    FN = 0
    TN = 0

    for i in range(len(y_pred)):
        if y_test[i] == 0 and y_pred[i] == 0:
            TN = TN + 1
        elif y_test[i] == 0 and y_pred[i] == 1:
            FP = FP + 1
        elif y_test[i] == 1 and y_pred[i] == 0:
            FN = FN + 1
        elif y_test[i] == 1 and y_pred[i] == 1:
            TP = TP + 1

    recall = recall_score(y_test, y_pred)
    precision = precision_score(y_test, y_pred)
    accuracy = accuracy_score(y_test, y_pred)
    f1_Score = f1_score(y_test, y_pred)
    return TN, FP, FN, TP, accuracy, f1_Score, recall, precision

# def export_eval_result(y_test, y_pred, test_result_path):
def export_eval_result(y_test, y_pred, result_dir, f_num=None):
    tn, fp, fn, tp, accuracy, f1, recall, precision = evaluate_data(y_test, y_pred)
    evaluation_result_df = pd.DataFrame({"TN" : [tn],
                                        "FP" : [fp],
                                        "FN" : [fn],
                                        "TP" : [tp],
                                        "Accuracy" : [accuracy],
                                        "F1_Score" : [f1],
                                        "Recall" : [recall],
                                        "precision" : [precision]})
    # evaluation_result_df.to_csv(test_result_path + "/evaluation_result.csv")
    evaluation_result_df.to_csv(f"{result_dir}/evaluation_result({f_num}).csv")


# def export_hit_eval_result(test_df, y_prob, test_result_path, model_name):
def export_hit_eval_result(test_df, result_dir, f_num=None):
    # thresholds = [0.3, 0.35, 0.4, 0.45, 0.5, 0.55, 0.6, 0.65, 0.7]
    thresholds = [0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9]
    total_Accuracy = []
    total_F1_Score = []
    total_Recall = []
    total_Precision = []
    total_inspection_persent = []
    total_minus_inspection_present = []
    total_hit_ratio = []
    total_TN = []
    total_FP = []
    total_FN = []
    total_TP = []

    for threshold in thresholds:
        thres_y_pred = []
        for i in range(len(test_df)):
            if test_df.y_prob[i] <= threshold:
                thres_y_pred.append(0)
            else:
                thres_y_pred.append(1)
        y_test = list(test_df.y_label)
        inspection_df = test_df.loc[threshold < test_df.y_prob]
        temp_df = test_df.drop(inspection_df.index, axis = 0)
        # if len(temp_df) == 0:
        #     continue
        inspection_len = len(inspection_df)
        try:
            inspection_persent = inspection_len/len(temp_df)
        except:
            inspection_persent = 0
        try:
            minus_inspection_persent = 1 - (inspection_len/len(temp_df))
        except:
            minus_inspection_persent = 1
        tn, fp, fn, tp, accuracy, f1, recall, precision = evaluate_data(y_test, thres_y_pred)
        hit_ratio = tp/(fn + tp)
        total_Accuracy.append(round(accuracy,4))
        total_F1_Score.append(round(f1, 4))
        total_Recall.append(round(recall, 4))
        total_Precision.append(round(precision, 4))
        total_inspection_persent.append(inspection_persent)
        total_hit_ratio.append(hit_ratio)
        total_TN.append(tn)
        total_FP.append(fp)
        total_FN.append(fn)
        total_TP.append(tp)            
        total_minus_inspection_present.append(minus_inspection_persent)
        
    hit_eval_df = pd.DataFrame({"threshold" : thresholds,
                            "TN" : total_TN,
                            "FP" : total_FP,
                            "FN" : total_FN,
                            "TP" : total_TP,
                            "hit_ratio" : total_hit_ratio,
                            "재검률" : total_inspection_persent,
                            "Accuracy" : total_Accuracy,
                            "F1_Score" : total_F1_Score,
                            "Recall" : total_Recall,
                            "precision" : total_Precision})
    # hit_eval_df.to_csv(test_result_path + "/hit_evaluation_result.csv", index = False)
    # plot_hit_ratio(hit_eval_df, total_minus_inspection_present, test_result_path, model_name)
    hit_eval_df.to_csv(f"{result_dir}/hit_evaluation_result({f_num}).csv", index = False)
    plot_hit_ratio(hit_eval_df, total_minus_inspection_present, result_dir, f_num)



# def export_test_result(test_df, y_pred, y_prob, test_result_path):
def export_test_result(test_df, y_pred, y_prob, result_dir, f_num=None):
    test_df['y_pred'] = y_pred
    test_df['y_prob'] = y_prob
    test_df.to_csv(f"{result_dir}/test_result({f_num}).csv")
    return test_df


# def plot_hit_ratio(df, total_minus_inspection_present, test_result_path, model_name):
def plot_hit_ratio(df, total_minus_inspection_present, result_dir, f_num):
    x = list(df["threshold"])
    y1 = list(df["hit_ratio"])
    y2 = total_minus_inspection_present

    fig, ax1 = plt.subplots()
    # plt.title(model_name)
    ax1.set_xlabel('Threshold')
    ax1.set_ylabel('hit_ratio')
    line1 = ax1.plot(x, y1, color = 'red', alpha = 0.5, label = "hit_ratio(%)", marker = "o")

    ax2 = ax1.twinx()
    ax2.set_ylabel('1 - inspection')
    line2 = ax2.plot(x, y2, color = 'blue', alpha = 0.5, label = "1 - inspection", marker = "o")

    lines = line1 + line2
    labels = [l.get_label() for l in lines]
    ax1.legend(lines, labels, loc='upper center')

    # plt.savefig(test_result_path + '/hit_ratio_image.jpg')
    plt.savefig(f'{result_dir}/hit_ratio_image({f_num}).jpg')
    plt.close()






def export_excel(model_list, result_dic):
    
    wb = openpyxl.Workbook()
    for i in range(len(model_list)):
        wb.create_sheet(model_list[i], i)
    wb.remove_sheet(wb["Sheet"])

    for ws in wb.worksheets:
        ws.merge_cells("A1:F1")
        ws["A1"] = ws.title + " 평가 결과"
        ws["A1"].font = Font(size = 24, bold = True)
        ws["A1"].alignment = Alignment(horizontal = "center", vertical = "center")

        ws.merge_cells("A3:F3")
        ws["A3"] = "cross validation result"
        ws["A3"].font = Font(size = 18)
        ws["A3"].alignment = Alignment(horizontal = "center", vertical = "center")

        column_temp = ["A", "B", "C", "D", "E", "F"]
        fold_val_title = ["fold", "loss", "accuracy", "f1-score", "precision", "recall"]
        for column, title in zip(column_temp, fold_val_title):
            ws[column + "4"] = title
            ws[column + "4"].alignment = Alignment(horizontal = "center", vertical = "center")
        
        temp_df = pd.read_csv(result_dic[ws.title] + "/" + ws.title + "/train_result/cross_val.csv")

        for xlsx_column, df_column in zip(["B", "C", "D", "E", "F"], [0, 1, 2, 3, 4]):
            for xlsx_low, df_low in zip([5, 6, 7, 8, 9], [0, 1, 2, 3, 4]):
                ws[xlsx_column + str(xlsx_low)] = temp_df.iloc[df_column][df_low]
                ws[xlsx_column + str(xlsx_low)].alignment = Alignment(horizontal = "center", vertical = "center")
        
        for i in range(5):
            ws["A" + str(i + 5)] = i + 1
            ws["A" + str(i + 5)].alignment = Alignment(horizontal = "center", vertical = "center")

        ws.merge_cells("A11:E11")
        ws["A11"] = "cross validation 평균 & 표준편차"
        ws["A11"].font = Font(size = 18)
        ws["A11"].alignment = Alignment(horizontal = "center", vertical = "center")

        column_temp_list = ["A", "B", "C", "D", "E"]
        title_temp_list = [['loss 평균', 'loss 표준편차', 'acc 평균', 'acc 표준편차', 'f1_score 평균'], ['f1_score 표준편차', 'recall 평균', 'recall 표준편차', 'precision 평균', 'precision 표준편차']]

        for i in range(len(title_temp_list)):
            for column, title in zip(column_temp_list, title_temp_list[i]):
                ws[column + str(12 + 2*i)] = title
                ws.column_dimensions[column].width = 16.5
                ws[column + str(12 + 2*i)].alignment = Alignment(horizontal = "center", vertical = "center")
        
        temp_df = pd.read_csv(result_dic[ws.title] + "/" + ws.title + "/train_result/" + 'result.csv')

        for xlsx_column, df_low in zip(["A", "B", "C", "D", "E"], [1, 2, 3, 4, 5]):
            ws[xlsx_column + "13"] = temp_df.iloc[0][df_low]
            ws[xlsx_column + "13"].alignment = Alignment(horizontal = "center", vertical = "center")
        
        for xlsx_column, df_low in zip(["A", "B", "C", "D", "E"], [6, 7, 8, 9, 10]):
            ws[xlsx_column + "15"] = temp_df.iloc[0][df_low]
            ws[xlsx_column + "15"].alignment = Alignment(horizontal = "center", vertical = "center")

        best_fold_num = temp_df.iloc[0][11]

        temp_df = pd.read_csv(result_dic[ws.title] + "/" + ws.title + "/test_result/hit_evaluation_result.csv")
        
        ws.merge_cells("A17:K17")
        ws["A17"] = "evaluation result"
        ws["A17"].font = Font(size = 18)
        ws["A17"].alignment = Alignment(horizontal = "center", vertical = "center")

        column_temp = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]
        evaluation_title = list(temp_df.columns)

        for column, title in zip(column_temp, evaluation_title):
            ws[column + "18"] = title
            ws.column_dimensions[column].width = 16.5
            ws[column + "18"].alignment = Alignment(horizontal = "center", vertical = "center")

        for xlsx_column, df_column in zip([19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29], [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]):
            for xlsx_low, df_low in zip(["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"], [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10]):
                ws[xlsx_low + str(xlsx_column)] = temp_df.iloc[df_column][df_low]
                ws[xlsx_low + str(xlsx_column)].alignment = Alignment(horizontal = "center", vertical = "center")

        image_path = result_dic[ws.title] + "/" + ws.title + "/test_result/hit_ratio_image.jpg"
        img = Image(image_path)
        img.height = 405.9
        img.width = 540.9
        ws.add_image(img, "G1")
    print(result_dic[model_list[0]] + "/" + model_list[0] + "/report.xlsx")
    wb.save(result_dic[model_list[0]] + "/" + model_list[0] +  "/report.xlsx")